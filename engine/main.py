# engine\main.py
import fitz
import io
import os
import platform
import subprocess
import string
import base64
import uuid
import qrcode
import sys
import webbrowser
import getpass
import shutil
import json
import secrets
import zipfile
import re
from datetime import datetime
from fastapi import FastAPI, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from pyzbar.pyzbar import decode
from typing import List, Optional
from PIL import Image

app = FastAPI(title="CoreKit Core Engine")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], 
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_ILLEGAL_CHARS = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_excel_val(val):
    if val is None or str(val).upper() == 'NULL':
        return None  
    return EXCEL_ILLEGAL_CHARS.sub('', str(val))

# ==========================================
# 1. STATE GLOBAL APLIKASI 
# ==========================================
class DocumentSession:
    def __init__(self, file_path=None, doc_bytes=None, is_pdf=True):
        if doc_bytes:
            self.doc = fitz.open(stream=doc_bytes, filetype="pdf")
        elif file_path:
            if is_pdf:
                with open(file_path, "rb") as f:
                    file_data = f.read()
                self.doc = fitz.open(stream=file_data, filetype="pdf")
            else:
                self.doc = fitz.open()
        else:
            self.doc = fitz.open()
            
        self.history = []
        self.redo_stack = []

    def save_snapshot(self):
        self.history.append(self.doc.tobytes())
        if len(self.history) > 10:
            self.history.pop(0)
        self.redo_stack.clear()

    def undo(self):
        if not self.history: return False
        self.redo_stack.append(self.doc.tobytes())
        old_bytes = self.history.pop()
        self.doc.close()
        self.doc = fitz.open(stream=old_bytes, filetype="pdf")
        return True

    def redo(self):
        if not self.redo_stack: return False
        self.history.append(self.doc.tobytes())
        new_bytes = self.redo_stack.pop()
        self.doc.close()
        self.doc = fitz.open(stream=new_bytes, filetype="pdf")
        return True

    def get_status(self):
        return {
            "can_undo": len(self.history) > 0,
            "can_redo": len(self.redo_stack) > 0
        }

active_sessions: dict[str, DocumentSession] = {}
doc_filepaths: dict[str, str] = {}
zip_sessions: dict[str, bytes] = {}
blob_sessions: dict[str, bytes] = {}


# ==========================================
# 2. PYDANTIC SCHEMAS 
# ==========================================
class FilePathReq(BaseModel): 
    path: str
    password: Optional[str] = None
class BaseDocReq(BaseModel): doc_id: str
class SaveReq(BaseDocReq): path: str
class SaveImageReq(BaseDocReq): path: str; img_format: str
class CloseReq(BaseDocReq): pass
class PageReq(BaseDocReq): page_num: int
class MoveToReq(BaseDocReq): page_num: int; target_page: int
class DeleteReq(BaseDocReq): mode: str; page_num: int = 0; pages: str = ""
class RotateReq(BaseDocReq): pages: str; angle: int
class SplitReq(BaseDocReq):
    mode: str
    pages: str = ""
    per_page: int = 1
class MergeReq(BaseModel): 
    files: List[str]
    passwords: Optional[dict] = None
    doc_id: Optional[str] = None
    insert_mode: str = "append" 
    insert_page: int = -1
class NumberingReq(BaseDocReq): 
    format: str = "full"
    style: str = "arabic"
    pages: str = "all"
    start_mode: str = "continue"
    start_at: str = "1"
    custom_prefix: str = ""  
    custom_suffix: str = ""  
    custom_divider: str = "dari"
    position: str = "bottom_right"
class SignReq(BaseDocReq): page_num: int; image_b64: str; norm_x: float; norm_y: float; norm_w: float; norm_h: float; remove_bg: bool
class AutoSignReq(BaseDocReq): pages: str; image_b64: str; norm_x: float; norm_y: float; norm_w: float; norm_h: float; remove_bg: bool
class TextReq(BaseDocReq): page_num: int; text: str; font_size: int; color_hex: str; norm_x: float; norm_y: float; norm_w: float; norm_h: float
class ProtectReq(BaseDocReq): password: str; output_path: str
class ProtectBatchItem(BaseModel):
    doc_id: str
    filename: Optional[str] = None
class ProtectBatchReq(BaseModel):
    files: List[ProtectBatchItem]
    password: str
class LockBatchItem(BaseModel):
    doc_id: str
    filename: Optional[str] = None  
class LockBatchReq(BaseModel):
    files: List[LockBatchItem]
class LockReq(BaseDocReq): output_path: str
class OpenFolderReq(BaseModel): path: str
class QRCodeReq(BaseModel):
    link: str
    with_logo: bool
class QRScanReq(BaseModel): image_b64: str
class OpenLinkReq(BaseModel): url: str
class FilesToPdfReq(BaseModel): files: List[str]; output_path: Optional[str] = None
class FileInfoReq(BaseModel): path: str; doc_id: Optional[str] = None
class CompressReq(BaseModel):
    doc_id: str
    mode: str
    password: Optional[str] = None
class SqlReadReq(BaseModel): path: str
class ExportDataReq(BaseModel): 
    output_path: str; table_name: str; columns: list; rows: list; format: str
    original_create: Optional[str] = None
    original_alter: Optional[List[str]] = None
    db_name: Optional[str] = None
class ImportDataReq(BaseModel): file_path: str; format: str
class TableData(BaseModel): 
    table_name: str; columns: list; rows: list
    original_create: Optional[str] = None
    original_alter: Optional[List[str]] = None
class ExportAllDataReq(BaseModel): 
    output_path: str; tables: List[TableData]; format: str
    db_name: Optional[str] = None


# ==========================================
# 3. HELPER FUNCTIONS
# ==========================================
def parse_page_string(page_str, max_pages):
    target_pages = set()
    if not page_str or page_str.lower() == "all": return set(range(max_pages))
    try:
        parts = page_str.split(',')
        for part in parts:
            part = part.strip()
            if not part: continue
            if '-' in part:
                start, end = map(int, part.split('-'))
                start, end = max(1, start), min(max_pages, end)
                for p in range(start, end + 1): target_pages.add(p - 1)
            else:
                p = int(part)
                if 1 <= p <= max_pages: target_pages.add(p - 1)
        return target_pages
    except:
        raise ValueError("Invalid page format")

def to_roman(n):
    if n <= 0: return str(n)
    val = [1000, 900, 500, 400, 100, 90, 50, 40, 10, 9, 5, 4, 1]
    syb = ["M", "CM", "D", "CD", "C", "XC", "L", "XL", "X", "IX", "V", "IV", "I"]
    roman_num = ''
    i = 0
    while n > 0:
        for _ in range(n // val[i]):
            roman_num += syb[i]
            n -= val[i]
        i += 1
    return roman_num

def to_alpha(n):
    if n <= 0: return str(n)
    result = ""
    while n > 0:
        n -= 1
        result = chr(65 + (n % 26)) + result
        n //= 26
    return result

def from_roman(s):
    roman_vals = {'I': 1, 'V': 5, 'X': 10, 'L': 50, 'C': 100, 'D': 500, 'M': 1000}
    s = s.upper().strip()
    total = 0
    prev_val = 0
    for char in reversed(s):
        if char not in roman_vals: return 1 # Fallback jika karakter tidak valid
        val = roman_vals[char]
        if val < prev_val:
            total -= val
        else:
            total += val
        prev_val = val
    return max(1, total)

def from_alpha(s):
    s = s.upper().strip()
    total = 0
    for char in s:
        if not 'A' <= char <= 'Z': return 1 # Fallback jika bukan abjad
        total = total * 26 + (ord(char) - 64)
    return max(1, total)

def parse_custom_start(val_str, style):
    val_str = str(val_str).strip()
    if not val_str: return 1
    
    # Jika pengguna menginput angka murni (misal: "5"), langsung gunakan
    if val_str.isdigit():
        return max(1, int(val_str))
        
    # Jika pengguna menginput huruf/romawi, terjemahkan sesuai style
    if style in ["alpha_lower", "alpha_upper"]:
        return from_alpha(val_str)
    elif style in ["roman_lower", "roman_upper"]:
        return from_roman(val_str)
        
    return 1 # Fallback terakhir

def format_number(num, style):
    if style == "roman_lower": return to_roman(num).lower()
    if style == "roman_upper": return to_roman(num)
    if style == "alpha_lower": return to_alpha(num).lower()
    if style == "alpha_upper": return to_alpha(num)
    return str(num)

def get_resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def record_audit(session: DocumentSession, action: str, details: str = ""):
    try:
        user = getpass.getuser()
        now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        log_entry = {"timestamp": now_str, "user": user, "action": action, "details": details}
        
        if session and session.doc:
            meta = session.doc.metadata or {}
            keywords = meta.get("keywords", "") or ""
            
            logs = []
            if "CoreKit_LOG_JSON:" in keywords:
                try:
                    json_str = keywords.split("CoreKit_LOG_JSON:")[1]
                    logs = json.loads(json_str)
                except:
                    pass
                    
            logs.insert(0, log_entry)
            
            clean_keywords = keywords.split("CoreKit_LOG_JSON:")[0]
            new_keywords = f"{clean_keywords}CoreKit_LOG_JSON:{json.dumps(logs)}"
            
            meta["keywords"] = new_keywords
            meta["modDate"] = fitz.get_pdf_now()
            session.doc.set_metadata(meta)
    except Exception as e:
        print(f"Failed to record audit log: {str(e)}")

# ==========================================
# 4. CORE ENDPOINTS 
# ==========================================
@app.get("/")
def read_root(): return {"status": "CoreKit Engine is Ready!"}

@app.post("/doc/open")
def open_pdf(data: FilePathReq):
    try:
        ext = data.path.lower().split('.')[-1]
        is_pdf = ext == 'pdf'

        session = DocumentSession(file_path=data.path, is_pdf=is_pdf)

        if not is_pdf:
            doc_id = str(uuid.uuid4())
            active_sessions[doc_id] = session
            doc_filepaths[doc_id] = data.path

            return {
                "status": "success",
                "doc_id": doc_id,
                "filename": os.path.basename(data.path),
                "total_pages": 0,
                "permissions": {},
                **session.get_status()
            }

        # Deteksi PDF Terproteksi (Password)
        if session.doc.needs_pass:
            if not data.password:
                return {"status": "needs_password"}
            if not session.doc.authenticate(data.password):
                return {"status": "wrong_password"}
                
        p = session.doc.permissions
        perms = {
            "commenting": bool(p & getattr(fitz, "PDF_PERM_ANNOTATE", 32)),
            "copying": bool(p & getattr(fitz, "PDF_PERM_COPY", 16)),
            "copying_accessibility": bool(p & getattr(fitz, "PDF_PERM_ACCESSIBILITY", 512)),
            "editing": bool(p & getattr(fitz, "PDF_PERM_MODIFY", 8)),
            "filling_forms": bool(p & getattr(fitz, "PDF_PERM_FORM", 256)),
            "printing": bool(p & getattr(fitz, "PDF_PERM_PRINT", 4)),
            "signing": bool(p & getattr(fitz, "PDF_PERM_MODIFY", 8))
        }

        doc_id = str(uuid.uuid4())
        active_sessions[doc_id] = session
        doc_filepaths[doc_id] = data.path
        
        return {
            "status": "success", 
            "doc_id": doc_id,
            "filename": os.path.basename(data.path),
            "total_pages": len(session.doc),
            "permissions": perms,
            **session.get_status()
        }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/doc/close")
def close_pdf(data: CloseReq):
    session = active_sessions.pop(data.doc_id, None)
    doc_filepaths.pop(data.doc_id, None)
    if session:
        session.history.clear()
        session.redo_stack.clear()
        session.doc.close()
    return {"status": "success"}

@app.post("/doc/save")
def save_pdf(data: SaveReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        session.doc.save(data.path, garbage=3, deflate=True)
        return {"status": "success", "path": data.path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.post("/doc/save_image")
def save_pdf_as_image(data: SaveImageReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        fmt = data.img_format.lower()
        if fmt not in ["jpg", "jpeg", "png"]: fmt = "jpg"
        doc = session.doc
        total_pages = len(doc)
        mat = fitz.Matrix(3.0, 3.0) 

        if total_pages == 1:
            page = doc[0]
            pix = page.get_pixmap(matrix=mat, alpha=False)
            with open(data.path, "wb") as f:
                f.write(pix.tobytes(fmt))
        else:
            with zipfile.ZipFile(data.path, "w", zipfile.ZIP_DEFLATED) as zipf:
                for i in range(total_pages):
                    page = doc[i]
                    pix = page.get_pixmap(matrix=mat, alpha=False)
                    zipf.writestr(f"Page_{i+1}.{fmt}", pix.tobytes(fmt))

        return {"status": "success", "path": data.path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/doc/download/{doc_id}")
def download_pdf_stream(doc_id: str, filename: str = "document.pdf"):
    session = active_sessions.get(doc_id)
    if not session: 
        raise HTTPException(status_code=400, detail="Invalid Document ID")
    
    try:
        # Ambil hasil manipulasi PDF dari memori menjadi bytes
        pdf_bytes = session.doc.tobytes(garbage=3, deflate=True)
        
        # Kirim kembali ke browser sebagai file attachment untuk di-download
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/doc/download_zip/{zip_id}")
def download_zip(zip_id: str, filename: str = "CoreKit-Split-Results.zip"):
    zip_bytes = zip_sessions.pop(zip_id, None)
    if zip_bytes is None:
        raise HTTPException(status_code=400, detail="Invalid Zip ID")
    return StreamingResponse(
        io.BytesIO(zip_bytes),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.get("/doc/download_blob/{blob_id}")
def download_blob(blob_id: str, filename: str = "document.pdf"):
    blob_bytes = blob_sessions.pop(blob_id, None)
    if blob_bytes is None:
        raise HTTPException(status_code=400, detail="Invalid Blob ID")
    return StreamingResponse(
        io.BytesIO(blob_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )

@app.post("/doc/undo")
def undo_action(data: BaseDocReq):
    session = active_sessions.get(data.doc_id)
    if session and session.undo(): return {"status": "success", "total_pages": len(session.doc), **session.get_status()}
    raise HTTPException(status_code=400, detail="Nothing to undo")

@app.post("/doc/redo")
def redo_action(data: BaseDocReq):
    session = active_sessions.get(data.doc_id)
    if session and session.redo(): return {"status": "success", "total_pages": len(session.doc), **session.get_status()}
    raise HTTPException(status_code=400, detail="Nothing to redo")

@app.get("/doc/render/{doc_id}/{page_num}")
def render_page(doc_id: str, page_num: int, zoom: float = 1.0):
    session = active_sessions.get(doc_id)
    if not session or page_num < 0 or page_num >= len(session.doc):
        raise HTTPException(status_code=404, detail="Page not found")
    page = session.doc[page_num]
    mat = fitz.Matrix(zoom * 2.0, zoom * 2.0)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    img_pil = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    img_pil = img_pil.resize((int(pix.width / 2.0), int(pix.height / 2.0)), Image.Resampling.LANCZOS)
    img_byte_arr = io.BytesIO()
    img_pil.save(img_byte_arr, format='WEBP', quality=85)
    img_byte_arr.seek(0)
    return StreamingResponse(img_byte_arr, media_type="image/webp")

# ==========================================
# 5. DOCUMENT CONVERSION
# ==========================================
@app.post("/tools/files_to_pdf")
def files_to_pdf(data: FilesToPdfReq):
    try:
        new_doc = fitz.open()
        for file_path in data.files:
            # [PERUBAHAN] Resolve doc_id ke path asli, sama seperti /tools/merge
            target_path = doc_filepaths.get(file_path, file_path)
            ext = target_path.lower().split('.')[-1]

            if ext in ['png', 'jpg', 'jpeg', 'webp', 'bmp']:
                img_doc = fitz.open(target_path)
                pdf_bytes = img_doc.convert_to_pdf()
                img_pdf = fitz.open("pdf", pdf_bytes)
                new_doc.insert_pdf(img_pdf)
                img_doc.close()
                img_pdf.close()
            elif ext == 'txt':
                with open(target_path, 'r', encoding='utf-8') as f:
                    text_content = f.read()
                page = new_doc.new_page(width=595, height=842)
                rect = fitz.Rect(50, 50, 545, 792)
                page.insert_textbox(rect, text_content, fontsize=11, fontname="helv")

        # [PERUBAHAN] Buat session in-memory, bukan simpan ke disk langsung
        doc_id = str(uuid.uuid4())
        session = DocumentSession(doc_bytes=new_doc.tobytes())
        new_doc.close()
        active_sessions[doc_id] = session
        doc_filepaths[doc_id] = "Converted_Document.pdf"
        record_audit(session, "Files to PDF", f"{len(data.files)} files converted")

        return {
            "status": "success",
            "doc_id": doc_id,
            "total_pages": len(session.doc),
            "filename": "Converted_Document.pdf",
            **session.get_status()
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ==========================================
# 6. PDF TOOLS
# ==========================================
@app.post("/tools/merge")
def merge_pdf(data: MergeReq):
    try:
        if data.doc_id and data.doc_id in active_sessions:
            session = active_sessions[data.doc_id]
            session.save_snapshot()
            start_idx = data.insert_page if data.insert_mode == "custom" and data.insert_page >= 0 else -1
            
            for f in data.files:
                # [PERUBAHAN] Ambil path asli dari memory jika 'f' adalah doc_id
                target_path = doc_filepaths.get(f, f)
                src_doc = fitz.open(target_path)
                
                if src_doc.needs_pass:
                    pwd = data.passwords.get(f, "") if data.passwords else ""
                    if not src_doc.authenticate(pwd):
                        raise ValueError(f"Password missing or incorrect for {f}")
                        
                if start_idx >= 0:
                    session.doc.insert_pdf(src_doc, start_at=start_idx)
                    start_idx += len(src_doc)
                else:
                    session.doc.insert_pdf(src_doc)
                src_doc.close()
                
            record_audit(session, "Merge PDF", f"{len(data.files)} files merged")
            return {
                "status": "success", 
                "doc_id": data.doc_id, 
                "total_pages": len(session.doc),
                "filename": doc_filepaths.get(data.doc_id, "Merged_Document.pdf"),
                **session.get_status()
            }
        else:
            new_doc = fitz.open()
            for f in data.files: 
                # [PERUBAHAN] Ambil path asli dari memory jika 'f' adalah doc_id
                target_path = doc_filepaths.get(f, f)
                src_doc = fitz.open(target_path)
                
                if src_doc.needs_pass:
                    pwd = data.passwords.get(f, "") if data.passwords else ""
                    if not src_doc.authenticate(pwd):
                        raise ValueError(f"Password missing or incorrect for {f}")
                new_doc.insert_pdf(src_doc)
                src_doc.close()
                
            doc_id = str(uuid.uuid4())
            session = DocumentSession(doc_bytes=new_doc.tobytes())
            new_doc.close()
            active_sessions[doc_id] = session
            doc_filepaths[doc_id] = "Merged_Document.pdf"
            record_audit(session, "Merge PDF", f"New document created from {len(data.files)} files")
            return {
                "status": "success", 
                "doc_id": doc_id, 
                "total_pages": len(session.doc),
                "filename": "Merged_Document.pdf",
                **session.get_status()
            }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

# Replace the whole /tools/split endpoint with:
@app.post("/tools/split")
def split_pdf(data: SplitReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        total = len(session.doc)

        if data.mode == "custom":
            targets = sorted(parse_page_string(data.pages, total))
            if not targets: raise ValueError("No valid pages selected")

            new_doc = fitz.open()
            for idx in targets:
                new_doc.insert_pdf(session.doc, from_page=idx, to_page=idx)

            doc_id = str(uuid.uuid4())
            new_session = DocumentSession(doc_bytes=new_doc.tobytes())
            new_doc.close()
            active_sessions[doc_id] = new_session
            filename = "CoreKit-Split-Custom.pdf"
            doc_filepaths[doc_id] = filename
            record_audit(new_session, "Split PDF (Custom Range)", f"Pages: {data.pages}")

            return {
                "status": "success",
                "mode": "custom",
                "doc_id": doc_id,
                "filename": filename,
                "total_pages": len(new_session.doc),
                "file_count": 1,
            }

        elif data.mode == "fixed":
            if data.per_page < 1: raise ValueError("Invalid page range size")

            parts = []
            for i in range(0, total, data.per_page):
                end = min(i + data.per_page, total)
                part_doc = fitz.open()
                part_doc.insert_pdf(session.doc, from_page=i, to_page=end - 1)
                parts.append(part_doc.tobytes())
                part_doc.close()

            if len(parts) == 1:
                doc_id = str(uuid.uuid4())
                new_session = DocumentSession(doc_bytes=parts[0])
                active_sessions[doc_id] = new_session
                filename = "CoreKit-Split-Part1.pdf"
                doc_filepaths[doc_id] = filename
                record_audit(new_session, "Split PDF (Fixed Range)", f"Per {data.per_page} pages")

                return {
                    "status": "success",
                    "mode": "fixed",
                    "doc_id": doc_id,
                    "filename": filename,
                    "total_pages": len(new_session.doc),
                    "file_count": 1,
                }

            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                for i, part_bytes in enumerate(parts):
                    zipf.writestr(f"CoreKit-Split-Part{i+1}.pdf", part_bytes)

            zip_id = str(uuid.uuid4())
            zip_sessions[zip_id] = zip_buffer.getvalue()
            record_audit(session, "Split PDF (Fixed Range)", f"Per {data.per_page} pages, {len(parts)} files")

            return {
                "status": "success",
                "mode": "fixed",
                "zip_id": zip_id,
                "filename": "CoreKit-Split-Results.zip",
                "file_count": len(parts),
            }
        else:
            raise ValueError("Invalid split mode")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/tools/rotate")
def rotate_pdf(data: RotateReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        targets = parse_page_string(data.pages, len(session.doc))
        if not targets: raise ValueError("No valid pages selected")
        session.save_snapshot()
        for idx in targets:
            session.doc[idx].set_rotation(session.doc[idx].rotation + data.angle)
        record_audit(session, "Rotate Page", f"Target: {data.pages}, Angle: {data.angle}°")
        return {"status": "success", **session.get_status()}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/page/move_up")
def move_page_up(data: PageReq):
    session = active_sessions.get(data.doc_id)
    if session and data.page_num > 0:
        session.save_snapshot()
        session.doc.move_page(data.page_num, data.page_num - 1)
        record_audit(session, "Move Page", f"Page {data.page_num + 1} moved up")
        return {"status": "success", **session.get_status()}
    raise HTTPException(status_code=400, detail="Invalid move")

@app.post("/page/move_down")
def move_page_down(data: PageReq):
    session = active_sessions.get(data.doc_id)
    if session and data.page_num < len(session.doc) - 1:
        session.save_snapshot()
        session.doc.move_page(data.page_num + 1, data.page_num)
        record_audit(session, "Move Page", f"Page {data.page_num + 1} moved down")
        return {"status": "success", **session.get_status()}
    raise HTTPException(status_code=400, detail="Invalid move")

@app.post("/page/move_to")
def move_page_to(data: MoveToReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    total_pages = len(session.doc)
    if data.page_num < 0 or data.page_num >= total_pages or data.target_page < 0 or data.target_page >= total_pages:
        raise HTTPException(status_code=400, detail="Target page out of bounds")
    try:
        session.save_snapshot()
        seq = list(range(total_pages))
        popped_page = seq.pop(data.page_num)
        seq.insert(data.target_page, popped_page)
        session.doc.select(seq)
        record_audit(session, "Move Page", f"Page {data.page_num + 1} moved to position {data.target_page + 1}")
        return {"status": "success", **session.get_status()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tools/numbering")
def add_page_numbers(data: NumberingReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        session.save_snapshot()
        doc = session.doc
        total_pages = len(doc)
        
        # Validasi target pages
        targets = sorted(list(parse_page_string(data.pages, total_pages)))
        if not targets: raise ValueError("No valid pages selected")
        
        total_str = format_number(total_pages, data.style)
        
        for seq_idx, page_idx in enumerate(targets):
            page = doc[page_idx]
            
            if data.start_mode == "start_at":
                parsed_start = parse_custom_start(data.start_at, data.style)
                current_num = parsed_start + seq_idx
            else:
                current_num = page_idx + 1
                
            curr_str = format_number(current_num, data.style)
            
            if data.format == "simple":
                text = curr_str
            elif data.format == "custom":
                prefix = f"{data.custom_prefix} " if data.custom_prefix.strip() else ""
                suffix = f" {data.custom_suffix}" if data.custom_suffix.strip() else ""
                div_val = data.custom_divider.strip()
                divider = f" {div_val} " if div_val else " "
                
                text = f"{prefix}{curr_str}{divider}{total_str}{suffix}"
            else:
                text = f"Halaman {curr_str} dari {total_str} halaman"
                
            text_width = fitz.get_text_length(text, fontname="helv", fontsize=10)
            pdf_w = page.rect.width
            pdf_h = page.rect.height
            
            # Margin aman dari tepi kertas
            margin_x = 30
            margin_y = 20
            
            # Logika kalkulasi posisi
            if data.position == "top_left":
                visual_x = margin_x
                visual_y = margin_y + 10 # +10 agar teks tidak terpotong ke atas
            elif data.position == "top_center":
                visual_x = (pdf_w - text_width) / 2
                visual_y = margin_y + 10
            elif data.position == "top_right":
                visual_x = pdf_w - text_width - margin_x
                visual_y = margin_y + 10
            elif data.position == "bottom_left":
                visual_x = margin_x
                visual_y = pdf_h - margin_y
            elif data.position == "bottom_center":
                visual_x = (pdf_w - text_width) / 2
                visual_y = pdf_h - margin_y
            else: # bottom_right (Default)
                visual_x = pdf_w - text_width - margin_x
                visual_y = pdf_h - margin_y

            p_visual = fitz.Point(visual_x, visual_y)
            p_internal = p_visual * page.derotation_matrix
            
            page.insert_text(p_internal, text, fontsize=10, fontname="helv", color=(0, 0, 0), rotate=page.rotation)
            
        record_audit(session, "Page Numbering", f"Format: {data.format}, Pos: {data.position}")
        return {"status": "success", **session.get_status()}
    except Exception as e:
        session.undo()
        raise HTTPException(status_code=500, detail=str(e))

def _process_signature_image(image_b64: str, remove_bg: bool) -> bytes:
    header, encoded = image_b64.split(",", 1)
    image_data = base64.b64decode(encoded)
    img = Image.open(io.BytesIO(image_data)).convert("RGBA")
    if remove_bg:
        datas = img.getdata()
        new_data = [(255, 255, 255, 0) if (item[0] > 240 and item[1] > 240 and item[2] > 240) else item for item in datas]
        img.putdata(new_data)
    img_byte_arr = io.BytesIO()
    img.save(img_byte_arr, format="PNG")
    return img_byte_arr.getvalue()

@app.post("/tools/sign")
def add_signature(data: SignReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        img_bytes = _process_signature_image(data.image_b64, data.remove_bg)
        session.save_snapshot()
        page = session.doc[data.page_num]
        rect_visual = fitz.Rect(data.norm_x * page.rect.width, data.norm_y * page.rect.height, (data.norm_x + data.norm_w) * page.rect.width, (data.norm_y + data.norm_h) * page.rect.height)
        page.insert_image(rect_visual * page.derotation_matrix, stream=img_bytes, overlay=True, rotate=page.rotation)
        record_audit(session, "Manual Signature", f"Page {data.page_num + 1}")
        return {"status": "success", **session.get_status()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tools/sign_auto")
def add_signature_auto(data: AutoSignReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        targets = parse_page_string(data.pages, len(session.doc))
        if not targets: raise ValueError("No valid pages selected")
        img_bytes = _process_signature_image(data.image_b64, data.remove_bg)
        session.save_snapshot()
        for idx in targets:
            page = session.doc[idx]
            rect_visual = fitz.Rect(data.norm_x * page.rect.width, data.norm_y * page.rect.height, (data.norm_x + data.norm_w) * page.rect.width, (data.norm_y + data.norm_h) * page.rect.height)
            page.insert_image(rect_visual * page.derotation_matrix, stream=img_bytes, overlay=True, rotate=page.rotation)
        record_audit(session, "Auto Signature", f"Target: {data.pages}")
        return {"status": "success", **session.get_status()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tools/text")
def add_text_to_pdf(data: TextReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        session.save_snapshot()
        page = session.doc[data.page_num]
        pdf_w = page.rect.width
        pdf_h = page.rect.height
        
        h = data.color_hex.lstrip('#')
        r, g, b = tuple(int(h[i:i+2], 16) / 255.0 for i in (0, 2, 4))
        
        visual_x = data.norm_x * pdf_w
        visual_y = data.norm_y * pdf_h
        
        pdf_fontsize = data.font_size * 0.75 
        baseline_offset = pdf_fontsize * 0.85 
        line_height = pdf_fontsize * 1.2 
        
        lines = data.text.split('\n')
        
        for i, line in enumerate(lines):
            current_visual_y = visual_y + baseline_offset + (i * line_height)
            p_visual = fitz.Point(visual_x, current_visual_y)
            p_internal = p_visual * page.derotation_matrix
            
            page.insert_text(p_internal, line, fontsize=pdf_fontsize, fontname="helv", color=(r, g, b), rotate=page.rotation)
            
        record_audit(session, "Tambah Teks", f"Halaman {data.page_num + 1}")
        return {"status": "success", **session.get_status()}
    except Exception as e:
        print(f"Error in /tools/text: {str(e)}") 
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tools/compress")
def compress_pdf(data: CompressReq):
    target_path = doc_filepaths.get(data.doc_id, data.doc_id)
    if not target_path or not os.path.exists(target_path):
        raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        doc = fitz.open(target_path)

        if doc.needs_pass:
            if not data.password or not doc.authenticate(data.password):
                raise ValueError("Password missing or incorrect for compression")

        if data.mode == "extreme": max_dim, jpeg_quality, garbage_lvl, subsampling = 720, 25, 4, 2
        elif data.mode == "recommended": max_dim, jpeg_quality, garbage_lvl, subsampling = 1200, 50, 4, 1
        else: max_dim, jpeg_quality, garbage_lvl, subsampling = 2000, 75, 3, 0  # "less"

        processed_xrefs, smask_dict = set(), {}
        for page_num in range(len(doc)):
            for img in doc[page_num].get_images(full=True):
                if img[1] > 0: smask_dict[img[0]] = img[1]

        smask_values = set(smask_dict.values())
        for page_num in range(len(doc)):
            page = doc[page_num]
            for img in page.get_images(full=True):
                xref = img[0]
                if xref in processed_xrefs or xref in smask_values or doc.xref_get_key(xref, "ImageMask")[1] == "true": continue
                processed_xrefs.add(xref)

                try:
                    image_info = doc.extract_image(xref)
                    orig_size = len(image_info["image"]) if image_info else doc.xref_length(xref)
                    img_pil = Image.open(io.BytesIO(fitz.Pixmap(doc, xref).tobytes("png")))
                    has_smask = xref in smask_dict

                    if not has_smask and img_pil.mode in ("RGBA", "LA", "P"):
                        img_pil = img_pil.convert("RGBA")
                        bg = Image.new("RGB", img_pil.size, (255, 255, 255))
                        bg.paste(img_pil, mask=img_pil.split()[3])
                        img_pil = bg
                    elif img_pil.mode not in ("RGB", "L"):
                        img_pil = img_pil.convert("RGB")

                    pdf_colorspace = "/DeviceGray" if img_pil.mode == "L" else "/DeviceRGB"
                    if img_pil.size[0] > max_dim or img_pil.size[1] > max_dim:
                        img_pil.thumbnail((max_dim, max_dim), Image.Resampling.LANCZOS)

                    img_byte_arr = io.BytesIO()
                    img_pil.save(img_byte_arr, format="JPEG", quality=jpeg_quality, optimize=True, subsampling=subsampling)
                    new_bytes = img_byte_arr.getvalue()

                    if len(new_bytes) < orig_size:
                        doc.update_stream(xref, new_bytes, compress=0)
                        doc.xref_set_key(xref, "Filter", "/DCTDecode")
                        doc.xref_set_key(xref, "ColorSpace", pdf_colorspace)
                        doc.xref_set_key(xref, "BitsPerComponent", "8")
                        doc.xref_set_key(xref, "Width", str(img_pil.size[0]))
                        doc.xref_set_key(xref, "Height", str(img_pil.size[1]))

                        keys_to_remove = ["Decode", "DecodeParms", "ColorTransform"]
                        if not has_smask: keys_to_remove.extend(["SMask", "Mask"])
                        for key in keys_to_remove:
                            if doc.xref_get_key(xref, key)[0] != "null": doc.xref_set_key(xref, key, "null")

                        if has_smask:
                            smask_xref = smask_dict[xref]
                            try:
                                smask_pil = Image.open(io.BytesIO(fitz.Pixmap(doc, smask_xref).tobytes("png"))).convert("L")
                                if smask_pil.size != img_pil.size: smask_pil = smask_pil.resize(img_pil.size, Image.Resampling.LANCZOS)
                                doc.update_stream(smask_xref, smask_pil.tobytes(), compress=1)
                                doc.xref_set_key(smask_xref, "Filter", "/FlateDecode")
                                doc.xref_set_key(smask_xref, "ColorSpace", "/DeviceGray")
                                doc.xref_set_key(smask_xref, "BitsPerComponent", "8")
                                doc.xref_set_key(smask_xref, "Width", str(img_pil.size[0]))
                                doc.xref_set_key(smask_xref, "Height", str(img_pil.size[1]))
                            except Exception: pass
                except Exception: continue

        original_size = os.path.getsize(target_path)
        compressed_bytes = doc.tobytes(garbage=garbage_lvl, deflate=True, clean=True)
        doc.close()

        # Jika hasil kompresi malah lebih besar, gunakan file asli sebagai fallback
        if len(compressed_bytes) >= original_size:
            with open(target_path, "rb") as f:
                compressed_bytes = f.read()

        result_doc_id = str(uuid.uuid4())
        new_session = DocumentSession(doc_bytes=compressed_bytes)
        active_sessions[result_doc_id] = new_session

        base_name = os.path.splitext(os.path.basename(target_path))[0]
        filename = f"{base_name}_compressed.pdf"
        doc_filepaths[result_doc_id] = filename
        record_audit(new_session, "Compress PDF", f"Mode: {data.mode}")

        return {
            "status": "success",
            "doc_id": result_doc_id,
            "filename": filename,
            "total_pages": len(new_session.doc),
            "original_size_bytes": original_size,
            "compressed_size_bytes": len(compressed_bytes),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tools/qrcode")
def generate_qrcode(data: QRCodeReq):
    try:
        if not data.link or not data.link.strip():
            raise ValueError("Link tidak boleh kosong")

        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_H if data.with_logo else qrcode.constants.ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        qr.add_data(data.link)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

        if data.with_logo:
            logo_path = get_resource_path(os.path.join("src", "assets", "bpom.png"))
            if not os.path.exists(logo_path):
                logo_path = os.path.join("..", "frontend", "public", "assets", "bpom.png")
            if os.path.exists(logo_path):
                logo = Image.open(logo_path).convert("RGBA")
                basewidth = int(img.size[0] / 4)
                wpercent = (basewidth / float(logo.size[0]))
                hsize = int((float(logo.size[1]) * float(wpercent)))
                logo = logo.resize((basewidth, hsize), Image.Resampling.LANCZOS)
                img.paste(logo, ((img.size[0] - logo.size[0]) // 2, (img.size[1] - logo.size[1]) // 2), mask=logo)

        img_byte_arr = io.BytesIO()
        img.save(img_byte_arr, format="PNG")
        img_b64 = base64.b64encode(img_byte_arr.getvalue()).decode("utf-8")

        return {
            "status": "success",
            "image_b64": f"data:image/png;base64,{img_b64}",
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tools/qrscan")
def scan_qrcode(data: QRScanReq):
    try:
        encoded = data.image_b64.split(",", 1)[1] if "," in data.image_b64 else data.image_b64
        decoded_objects = decode(Image.open(io.BytesIO(base64.b64decode(encoded))))
        if not decoded_objects: raise ValueError("QR Code not detected in image.")
        return {"status": "success", "data": decoded_objects[0].data.decode("utf-8")}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/security/protect")
def protect_pdf(data: ProtectReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        session.doc.save(data.output_path, encryption=fitz.PDF_ENCRYPT_AES_256, user_pw=data.password, owner_pw=data.password)
        record_audit(session, "Password Protection")
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/security/protect_batch")
def protect_pdf_batch(data: ProtectBatchReq):
    try:
        if not data.files:
            raise ValueError("Tidak ada file yang dipilih")
        if not data.password:
            raise ValueError("Password tidak boleh kosong")

        results = []
        for item in data.files:
            session = active_sessions.get(item.doc_id)
            if not session:
                raise ValueError(f"Invalid Document ID: {item.doc_id}")

            encrypted_bytes = session.doc.tobytes(
                encryption=fitz.PDF_ENCRYPT_AES_256,
                user_pw=data.password,
                owner_pw=data.password,
                garbage=3,
                deflate=True,
            )
            record_audit(session, "Password Protection")

            base_name = item.filename or doc_filepaths.get(item.doc_id, "Protected_Document.pdf")
            if not base_name.lower().endswith(".pdf"):
                base_name += ".pdf"
            results.append((base_name, encrypted_bytes))

        if len(results) == 1:
            filename, file_bytes = results[0]
            blob_id = str(uuid.uuid4())
            blob_sessions[blob_id] = file_bytes
            return {
                "status": "success",
                "mode": "single",
                "blob_id": blob_id,
                "filename": filename,
                "file_count": 1,
            }
        else:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                used_names = set()
                for filename, file_bytes in results:
                    final_name = filename
                    counter = 1
                    while final_name in used_names:
                        name_part, ext = os.path.splitext(filename)
                        final_name = f"{name_part}_{counter}{ext}"
                        counter += 1
                    used_names.add(final_name)
                    zipf.writestr(final_name, file_bytes)

            zip_id = str(uuid.uuid4())
            zip_sessions[zip_id] = zip_buffer.getvalue()
            return {
                "status": "success",
                "mode": "zip",
                "zip_id": zip_id,
                "filename": "CoreKit-Protected-Results.zip",
                "file_count": len(results),
            }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/security/lock_batch")
def lock_pdf_batch(data: LockBatchReq):
    try:
        if not data.files:
            raise ValueError("Tidak ada file yang dipilih")

        # Permission yang diizinkan: Printing, Content Copying, Copying for Accessibility.
        # Yang dibatasi: Commenting, Editing file content, Filling form fields, Signing.
        perms = fitz.PDF_PERM_PRINT | fitz.PDF_PERM_COPY | fitz.PDF_PERM_ACCESSIBILITY

        results = []
        for item in data.files:
            session = active_sessions.get(item.doc_id)
            if not session:
                raise ValueError(f"Invalid Document ID: {item.doc_id}")

            owner_pw = ''.join(secrets.choice(string.ascii_letters + string.digits) for i in range(16))
            locked_bytes = session.doc.tobytes(
                encryption=fitz.PDF_ENCRYPT_AES_256,
                owner_pw=owner_pw,
                user_pw="",
                permissions=perms,
                garbage=3,
                deflate=True,
            )
            record_audit(session, "Lock Document (Read-Only)")

            base_name = item.filename or doc_filepaths.get(item.doc_id, "Locked_Document.pdf")
            if not base_name.lower().endswith(".pdf"):
                base_name += ".pdf"
            results.append((base_name, locked_bytes))

        if len(results) == 1:
            filename, file_bytes = results[0]
            blob_id = str(uuid.uuid4())
            blob_sessions[blob_id] = file_bytes
            return {
                "status": "success",
                "mode": "single",
                "blob_id": blob_id,
                "filename": filename,
                "file_count": 1,
            }
        else:
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zipf:
                used_names = set()
                for filename, file_bytes in results:
                    final_name = filename
                    counter = 1
                    while final_name in used_names:
                        name_part, ext = os.path.splitext(filename)
                        final_name = f"{name_part}_{counter}{ext}"
                        counter += 1
                    used_names.add(final_name)
                    zipf.writestr(final_name, file_bytes)

            zip_id = str(uuid.uuid4())
            zip_sessions[zip_id] = zip_buffer.getvalue()
            return {
                "status": "success",
                "mode": "zip",
                "zip_id": zip_id,
                "filename": "CoreKit-Locked-Results.zip",
                "file_count": len(results),
            }
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/security/lock")
def lock_pdf(data: LockReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        owner_pw = ''.join(secrets.choice(string.ascii_letters + string.digits) for i in range(16))
        perms = fitz.PDF_PERM_PRINT | fitz.PDF_PERM_COPY | fitz.PDF_PERM_ACCESSIBILITY
        session.doc.save(data.output_path, encryption=fitz.PDF_ENCRYPT_AES_256, owner_pw=owner_pw, user_pw="", permissions=perms)
        record_audit(session, "Lock Document (Read-Only)")
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

def parse_sql_values(row_str: str) -> list:
    values, current_val, in_string, quote_char, escape_next, parens_depth = [], [], False, '', False, 0
    for char in row_str:
        if escape_next:
            current_val.append(char)
            escape_next = False
        elif char == '\\':
            current_val.append(char)
            escape_next = True
        elif in_string:
            current_val.append(char)
            if char == quote_char: in_string = False
        else:
            if char in ("'", '"'):
                in_string, quote_char = True, char
                current_val.append(char)
            elif char == '(':
                parens_depth += 1
                current_val.append(char)
            elif char == ')':
                parens_depth -= 1
                current_val.append(char)
            elif char == ',' and parens_depth == 0:
                val = ''.join(current_val).strip()
                if len(val) >= 2 and val[0] in ("'", '"') and val[-1] == val[0]: val = val[1:-1]
                values.append(val)
                current_val = []
            else:
                current_val.append(char)
    val = ''.join(current_val).strip()
    if len(val) >= 2 and val[0] in ("'", '"') and val[-1] == val[0]: val = val[1:-1]
    values.append(val)
    return values

@app.post("/tools/read_data_structure")
def read_data_structure(data: SqlReadReq):
    try:
        ext = data.path.split('.')[-1].lower()
        tables_dict = {}
        db_name = None

        if ext in ['xlsx', 'xls']:
            try:
                import openpyxl
            except ImportError:
                raise ValueError("Library 'openpyxl' tidak ditemukan.")
            
            wb = openpyxl.load_workbook(data.path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)
                try: headers = next(rows_iter)
                except StopIteration: continue
                
                if not headers or all(h is None for h in headers): continue
                columns = [str(h) if h is not None else f"Col_{i}" for i, h in enumerate(headers)]
                parsed_rows = []
                for row in rows_iter:
                    if all(cell is None or str(cell).strip() == "" for cell in row): continue
                    parsed_rows.append([str(cell) if cell is not None else 'NULL' for cell in row])
                    
                tables_dict[sheet_name] = {"name": sheet_name, "columns": columns, "rows": parsed_rows}
        else:
            with open(data.path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()

            # Tangkap nama Database
            db_match = re.search(r'CREATE\s+DATABASE\s+[`"\'\[]?([a-zA-Z0-9_]+)[`"\'\]]?', content, re.IGNORECASE)
            if db_match: db_name = db_match.group(1)

            # Tangkap struktur utuh CREATE TABLE
            create_pattern = re.compile(r'CREATE\s+TABLE\s+[`"\'\[]?([a-zA-Z0-9_]+)[`"\'\]]?\s*\((.*?)\);', re.IGNORECASE | re.DOTALL)
            for match in create_pattern.finditer(content):
                t_name = match.group(1)
                full_create = match.group(0)
                col_str = match.group(2)
                
                raw_lines = [l.strip() for l in col_str.split('\n') if l.strip()]
                columns = []
                for line in raw_lines:
                    if line.upper().startswith(('PRIMARY', 'UNIQUE', 'KEY', 'CONSTRAINT', 'FOREIGN')): continue
                    col_match = re.match(r'[`"\'\[]?([a-zA-Z0-9_]+)[`"\'\]]?', line)
                    if col_match: columns.append(col_match.group(1))
                
                tables_dict[t_name] = {"name": t_name, "columns": columns, "rows": [], "original_create": full_create, "original_alter": []}

            # Tangkap struktur utuh ALTER TABLE
            alter_pattern = re.compile(r'ALTER\s+TABLE\s+[`"\'\[]?([a-zA-Z0-9_]+)[`"\'\]]?\s+ADD\s+[^;]+;', re.IGNORECASE | re.DOTALL)
            for match in alter_pattern.finditer(content):
                t_name = match.group(1)
                if t_name in tables_dict:
                    tables_dict[t_name]["original_alter"].append(match.group(0))

            # Tangkap insert records
            insert_pattern = re.compile(r'INSERT\s+INTO\s+[`"\'\[]?([a-zA-Z0-9_]+)[`"\'\]]?(?:\s*\([^)]+\))?\s+VALUES\s*(.*?);', re.IGNORECASE | re.DOTALL)
            for match in insert_pattern.finditer(content):
                t_name = match.group(1)
                values_str = match.group(2).strip()
                if t_name in tables_dict:
                    raw_tuples = re.split(r'\)\s*,\s*\(', values_str)
                    if raw_tuples:
                        raw_tuples[0] = re.sub(r'^\s*\(', '', raw_tuples[0])
                        raw_tuples[-1] = re.sub(r'\)\s*$', '', raw_tuples[-1])
                        
                    parsed_rows = []
                    # Ambil jumlah kolom yang valid dari hasil bacaan CREATE TABLE
                    expected_cols = len(tables_dict[t_name]["columns"]) 
                    
                    for rt in raw_tuples: 
                        row_vals = parse_sql_values(rt)
                        if expected_cols > 0 and len(row_vals) != expected_cols:
                            continue 
                            
                        parsed_rows.append(row_vals)
                    tables_dict[t_name]["rows"].extend(parsed_rows)

        return {"status": "success", "tables": list(tables_dict.values()), "filename": os.path.basename(data.path), "db_name": db_name}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/tools/export_data")
def export_data(data: ExportDataReq):
    try:
        if data.format == "sql":
            with open(data.output_path, 'w', encoding='utf-8') as f:
                if data.db_name:
                    f.write(f"create database {data.db_name};\n")
                    f.write(f"use {data.db_name};\n")
                
                if data.original_create:
                    f.write(data.original_create + "\n")
                else:
                    f.write(f"CREATE TABLE `{data.table_name}` (\n")
                    cols_def = [f"`{c}` TEXT DEFAULT NULL" for c in data.columns]
                    f.write(",\n".join(cols_def))
                    f.write(");\n")

                if data.original_alter:
                    for alt in data.original_alter: 
                        f.write(alt + "\n")

                if data.rows:
                    f.write("\n") # 1 baris kosong wajib sebelum blok INSERT
                    
                    # Membatasi 60 baris per query sesuai format sistem native
                    chunk_size = 60
                    for i in range(0, len(data.rows), chunk_size):
                        chunk = data.rows[i : i + chunk_size]
                        
                        if i > 0:
                            f.write("\n") # 1 baris kosong antar potongan INSERT
                            
                        f.write(f"INSERT INTO {data.table_name} VALUES ")
                        values_blocks = []
                        for row in chunk:
                            safe_vals = []
                            for v in row:
                                if v is None or str(v).upper() == 'NULL': 
                                    safe_vals.append('NULL')
                                else:
                                    val_esc = str(v).replace('"', '\\"')
                                    safe_vals.append(f'"{val_esc}"')
                            values_blocks.append(f"({','.join(safe_vals)})") 
                        f.write(",\n".join(values_blocks) + ";\n")
                    
        elif data.format == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = data.table_name[:31] 
            ws.append(data.columns)
            for row in data.rows: ws.append([clean_excel_val(v) for v in row])
            wb.save(data.output_path)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/tools/export_all_data")
def export_all_data(data: ExportAllDataReq):
    try:
        if data.format == "sql":
            with open(data.output_path, 'w', encoding='utf-8') as f:
                if data.db_name:
                    f.write(f"create database {data.db_name};\n")
                    f.write(f"use {data.db_name};\n")

                for idx, table in enumerate(data.tables):
                    if table.original_create:
                        f.write(table.original_create + "\n")
                    else:
                        f.write(f"CREATE TABLE `{table.table_name}` (\n")
                        cols_def = [f"`{c}` TEXT DEFAULT NULL" for c in table.columns]
                        f.write(",\n".join(cols_def))
                        f.write(");\n")
                    
                    if table.original_alter:
                        for alt in table.original_alter: 
                            f.write(alt + "\n")
                    
                    if table.rows:
                        f.write("\n") # 1 baris kosong wajib sebelum blok INSERT
                        
                        # Membatasi 60 baris per query sesuai format sistem native
                        chunk_size = 60
                        for i in range(0, len(table.rows), chunk_size):
                            chunk = table.rows[i : i + chunk_size]
                            
                            if i > 0:
                                f.write("\n") # 1 baris kosong antar potongan INSERT
                                
                            f.write(f"INSERT INTO {table.table_name} VALUES ")
                            blocks = []
                            for row in chunk:
                                safe_vals = []
                                for v in row:
                                    if v is None or str(v).upper() == 'NULL': 
                                        safe_vals.append('NULL')
                                    else:
                                        val_esc = str(v).replace('"', '\\"')
                                        safe_vals.append(f'"{val_esc}"')
                                blocks.append(f"({','.join(safe_vals)})")
                            f.write(",\n".join(blocks) + ";\n")
                    
                    # Logika Pemisah Antar Tabel yang Akurat & Dinamis
                    if idx < len(data.tables) - 1:
                        if table.rows:
                            f.write("\n")     # Jika tabel ada data, diakhiri 1 baris kosong
                        elif table.original_alter:
                            f.write("\n")     # Jika tabel ada alter, diakhiri 1 baris kosong
                        else:
                            f.write("\n\n")   # Jika hanya CREATE TABLE, diakhiri 2 baris kosong
                    
        elif data.format == "excel":
            import openpyxl
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            for table in data.tables:
                ws = wb.create_sheet(title=table.table_name[:31])
                ws.append(table.columns)          
                for row in table.rows: ws.append([clean_excel_val(v) for v in row])
            if len(wb.sheetnames) > 1: wb.remove(default_sheet)
            wb.save(data.output_path)

        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/system/open_folder")
def open_folder(data: OpenFolderReq):
    try:
        target_path = os.path.normpath(data.path)
        if platform.system() == "Windows":
            if os.path.isfile(target_path): subprocess.run(['explorer', '/select,', target_path])
            else: os.startfile(os.path.dirname(target_path) if not os.path.isdir(target_path) else target_path)
        elif platform.system() == "Darwin": subprocess.call(["open", "-R", target_path]) 
        else: 
            folder_path = os.path.dirname(target_path) if os.path.isfile(target_path) else target_path
            subprocess.call(["xdg-open", folder_path])
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/page/delete")
def delete_page(data: DeleteReq):
    session = active_sessions.get(data.doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        session.save_snapshot()
        if data.mode == "range":
            targets = parse_page_string(data.pages, len(session.doc))
            if not targets: raise ValueError("No valid pages selected")
            if len(targets) >= len(session.doc): raise ValueError("Cannot delete all pages in the document")
            for idx in sorted(list(targets), reverse=True): session.doc.delete_page(idx)
            record_audit(session, "Delete Pages", f"Pages {data.pages} deleted")
        else:
            if len(session.doc) <= 1: raise ValueError("Cannot delete last page")
            session.doc.delete_page(data.page_num)
            record_audit(session, "Delete Page", f"Page {data.page_num + 1} deleted")
        return {"status": "success", "total_pages": len(session.doc), **session.get_status()}
    except Exception as e:
        session.undo()
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/system/file_info")
def get_file_info(data: FileInfoReq):
    try:
        return {"status": "success", "size_bytes": os.path.getsize(data.path)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/system/open_link")
def open_link(data: OpenLinkReq):
    try:
        webbrowser.open(data.url)
        return {"status": "success"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/doc/audit_log/{doc_id}")
def get_document_audit_log(doc_id: str):
    session = active_sessions.get(doc_id)
    if not session: raise HTTPException(status_code=400, detail="Invalid Document ID")
    try:
        meta = session.doc.metadata or {}
        keywords = meta.get("keywords", "") or ""
        logs = []
        if "CoreKit_LOG_JSON:" in keywords: logs = json.loads(keywords.split("CoreKit_LOG_JSON:")[1])
        return {"status": "success", "data": logs}
    except Exception as e:
        return {"status": "success", "data": []}

@app.post("/system/shutdown")
def shutdown_engine():
    os._exit(0)

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)